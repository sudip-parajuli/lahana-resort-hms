"""
SIA HMS — Excel Report Builders
Uses openpyxl to generate beautifully formatted spreadsheet reports.
"""

import openpyxl
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from apps.staff.models import StaffMember
from apps.hr.models import Attendance
from apps.inventory.models import InventoryItem, StockMovement
from apps.payroll.models import PayrollEntry
from apps.billing.models import InvoiceItem
from django.db.models import Sum
from datetime import timedelta

# Shared Styling Constants
FONT_TITLE = Font(name="Segoe UI", size=16, bold=True, color="1E293B")
FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="64748B")
FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Segoe UI", size=10, color="334155")
FONT_TOTAL = Font(name="Segoe UI", size=11, bold=True, color="0F172A")

FILL_HEADER = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
FILL_ZEBRA = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
FILL_ACCENT = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
BORDER_TOTAL = Border(
    top=Side(style="thin", color="1E293B"),
    bottom=Side(style="double", color="1E293B")
)


def format_ws(ws, title, subtitle):
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_LEFT
    
    ws.merge_cells("A2:H2")
    ws["A2"] = subtitle
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = ALIGN_LEFT
    
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10  # empty row


def auto_fit_cols(ws, min_col=1, max_col=20):
    for col in ws.iter_cols(min_col=min_col, max_col=max_col):
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def build_attendance_summary_excel(start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Summary"
    
    format_ws(ws, "Staff Attendance Summary", f"Date range: {start_date} to {end_date}")
    
    # Header Columns
    headers = ["Employee", "Department"]
    date_cols = []
    curr = start_date
    while curr <= end_date:
        headers.append(curr.strftime("%d"))
        date_cols.append(curr)
        curr += timedelta(days=1)
    
    headers.extend(["Present", "Absent", "Leave"])
    
    # Write Header Row
    ws.row_dimensions[4].height = 25
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER if col_idx > 2 else ALIGN_LEFT
        cell.border = BORDER_THIN

    # Fetch Staff and Attendance
    staff_members = StaffMember.objects.all().order_by("user__first_name")
    row_idx = 5
    
    for staff in staff_members:
        ws.row_dimensions[row_idx].height = 20
        # Basic Info
        ws.cell(row=row_idx, column=1, value=staff.user.get_full_name() or staff.user.email).font = FONT_BODY
        ws.cell(row=row_idx, column=2, value=staff.department.name if staff.department else "-").font = FONT_BODY
        
        # Zebra striping
        fill = FILL_ZEBRA if row_idx % 2 == 0 else PatternFill(fill_type=None)
        
        p_count, a_count, l_count = 0, 0, 0
        
        col_idx = 3
        for d in date_cols:
            att = Attendance.objects.filter(staff=staff, date=d).first()
            status_val = "-"
            if att:
                if att.status == "present":
                    status_val = "P"
                    p_count += 1
                elif att.status == "absent":
                    status_val = "A"
                    a_count += 1
                elif att.status == "leave":
                    status_val = "L"
                    l_count += 1
            
            c = ws.cell(row=row_idx, column=col_idx, value=status_val)
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER
            c.border = BORDER_THIN
            if fill.fill_type:
                c.fill = fill
            col_idx += 1
            
        # Write counts
        for idx, count_val in enumerate([p_count, a_count, l_count], col_idx):
            c = ws.cell(row=row_idx, column=idx, value=count_val)
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER
            c.border = BORDER_THIN
            if fill.fill_type:
                c.fill = fill
                
        row_idx += 1
        
    auto_fit_cols(ws, 1, len(headers))
    
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def build_inventory_consumption_excel(start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Consumption"
    
    format_ws(ws, "Inventory Consumption Report", f"Date range: {start_date} to {end_date}")
    
    headers = ["Item Name", "Category", "SKU", "Unit", "Initial Stock", "Received (GRN)", "Consumed", "Adjusted", "Final Stock"]
    ws.row_dimensions[4].height = 25
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_LEFT if col_idx <= 4 else ALIGN_RIGHT
        cell.border = BORDER_THIN

    items = InventoryItem.objects.all().order_by("name")
    row_idx = 5
    
    for item in items:
        ws.row_dimensions[row_idx].height = 20
        # Calculate movements
        movements_after_start = StockMovement.objects.filter(item=item, created_at__date__gte=start_date)
        movements_in_range = movements_after_start.filter(created_at__date__lte=end_date)
        
        received = movements_in_range.filter(movement_type="purchase").aggregate(Sum("quantity"))["quantity__sum"] or 0
        consumed = movements_in_range.filter(movement_type="consumption").aggregate(Sum("quantity"))["quantity__sum"] or 0
        adjusted = movements_in_range.filter(movement_type="adjustment").aggregate(Sum("quantity"))["quantity__sum"] or 0
        
        current_stock = item.current_stock
        initial = current_stock
        
        ws.cell(row=row_idx, column=1, value=item.name).font = FONT_BODY
        ws.cell(row=row_idx, column=2, value=item.category.name if item.category else "-").font = FONT_BODY
        ws.cell(row=row_idx, column=3, value=item.sku or "-").font = FONT_BODY
        ws.cell(row=row_idx, column=4, value=item.unit or "-").font = FONT_BODY
        
        ws.cell(row=row_idx, column=5, value=float(initial)).font = FONT_BODY
        ws.cell(row=row_idx, column=6, value=float(received)).font = FONT_BODY
        ws.cell(row=row_idx, column=7, value=float(consumed)).font = FONT_BODY
        ws.cell(row=row_idx, column=8, value=float(adjusted)).font = FONT_BODY
        ws.cell(row=row_idx, column=9, value=float(item.current_stock)).font = FONT_BODY
        
        fill = FILL_ZEBRA if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for c in range(1, 10):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = BORDER_THIN
            if fill.fill_type:
                cell.fill = fill
            if c > 4:
                cell.alignment = ALIGN_RIGHT
                cell.number_format = "#,##0.00"
                
        row_idx += 1
        
    auto_fit_cols(ws, 1, len(headers))
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def build_profit_loss_excel(start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L Summary"
    
    format_ws(ws, "Profit & Loss Summary Statement", f"For period: {start_date} to {end_date}")
    
    room_rev = InvoiceItem.objects.filter(
        item_type="room_charge",
        invoice__payments__paid_at__date__range=(start_date, end_date)
    ).aggregate(Sum("amount"))["amount__sum"] or 0
    
    fb_rev = InvoiceItem.objects.filter(
        item_type="pos_charge",
        invoice__payments__paid_at__date__range=(start_date, end_date)
    ).aggregate(Sum("amount"))["amount__sum"] or 0
    
    other_rev = InvoiceItem.objects.filter(
        item_type="extra_charge",
        invoice__payments__paid_at__date__range=(start_date, end_date)
    ).aggregate(Sum("amount"))["amount__sum"] or 0
    
    gross_revenue = room_rev + fb_rev + other_rev
    
    from apps.inventory.models import PurchaseOrder
    po_expenses = PurchaseOrder.objects.filter(
        created_at__date__range=(start_date, end_date),
        status="received"
    ).aggregate(Sum("total_amount"))["total_amount__sum"] or 0
    
    payroll_salaries = PayrollEntry.objects.filter(
        period__created_at__date__range=(start_date, end_date)
    ).aggregate(Sum("basic_salary"), Sum("overtime_amount"), Sum("ssf_employer"))
    
    base_sal = payroll_salaries["basic_salary__sum"] or 0
    ot_sal = payroll_salaries["overtime_amount__sum"] or 0
    ssf_employer = payroll_salaries["ssf_employer__sum"] or 0
    
    total_expenses = po_expenses + base_sal + ot_sal + ssf_employer
    net_income = gross_revenue - total_expenses
    
    lines = [
        ("REVENUES", "", True),
        ("  Room Bookings Revenue", float(room_rev), False),
        ("  Food & Beverage POS Revenue", float(fb_rev), False),
        ("  Extra Service Charges Revenue", float(other_rev), False),
        ("TOTAL REVENUE", float(gross_revenue), True),
        ("", "", False),
        ("OPERATING EXPENSES", "", True),
        ("  Inventory Purchases (GRN)", float(po_expenses), False),
        ("  Staff Base Monthly Payroll", float(base_sal), False),
        ("  Staff Overtime Allocations", float(ot_sal), False),
        ("  Employer SSF Pension Contributions (20%)", float(ssf_employer), False),
        ("TOTAL OPERATING EXPENSES", float(total_expenses), True),
        ("", "", False),
        ("NET OPERATING INCOME", float(net_income), True),
    ]
    
    row_idx = 5
    for desc, val, bold in lines:
        ws.row_dimensions[row_idx].height = 20
        ws.cell(row=row_idx, column=1, value=desc).font = Font(name="Segoe UI", size=10, bold=bold)
        
        c_val = ws.cell(row=row_idx, column=2, value=val)
        c_val.font = Font(name="Segoe UI", size=10, bold=bold)
        if val != "":
            c_val.number_format = "#,##0.00"
            c_val.alignment = ALIGN_RIGHT
            
        if bold:
            ws.cell(row=row_idx, column=1).fill = FILL_ACCENT
            c_val.fill = FILL_ACCENT
            if desc in ["TOTAL REVENUE", "TOTAL OPERATING EXPENSES", "NET OPERATING INCOME"]:
                ws.cell(row=row_idx, column=1).border = BORDER_TOTAL
                c_val.border = BORDER_TOTAL
                
        row_idx += 1
        
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def build_payroll_summary_excel(start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Summary"
    
    format_ws(ws, "Monthly Payroll Reconciliation Ledger", f"Cycles ending: {start_date} to {end_date}")
    
    headers = [
        "Employee Name", "Department", "Base Salary", "Worked Days",
        "Overtime Pay", "Allowances", "SSF Employee (11%)",
        "Income Tax (TDS)", "Net Salary Paid"
    ]
    ws.row_dimensions[4].height = 25
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_LEFT if col_idx <= 2 else ALIGN_RIGHT
        cell.border = BORDER_THIN
        
    entries = PayrollEntry.objects.filter(
        period__created_at__date__range=(start_date, end_date)
    ).order_by("staff__user__first_name")
    
    row_idx = 5
    for entry in entries:
        ws.row_dimensions[row_idx].height = 20
        ws.cell(row=row_idx, column=1, value=entry.staff.user.get_full_name() or entry.staff.user.email).font = FONT_BODY
        ws.cell(row=row_idx, column=2, value=entry.staff.department.name if entry.staff.department else "-").font = FONT_BODY
        ws.cell(row=row_idx, column=3, value=float(entry.basic_salary)).font = FONT_BODY
        ws.cell(row=row_idx, column=4, value=entry.present_days).font = FONT_BODY
        ws.cell(row=row_idx, column=5, value=float(entry.overtime_amount)).font = FONT_BODY
        
        allowance_total = sum(x.get("amount", 0) for x in (entry.allowances or []))
        ws.cell(row=row_idx, column=6, value=float(allowance_total)).font = FONT_BODY
        ws.cell(row=row_idx, column=7, value=float(entry.ssf_employee)).font = FONT_BODY
        ws.cell(row=row_idx, column=8, value=float(entry.income_tax)).font = FONT_BODY
        ws.cell(row=row_idx, column=9, value=float(entry.net_salary)).font = FONT_BODY
        
        fill = FILL_ZEBRA if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for c in range(1, 10):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = BORDER_THIN
            if fill.fill_type:
                cell.fill = fill
            if c >= 3 and c != 4:
                cell.alignment = ALIGN_RIGHT
                cell.number_format = "#,##0.00"
            elif c == 4:
                cell.alignment = ALIGN_CENTER
                
        row_idx += 1
        
    auto_fit_cols(ws, 1, len(headers))
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
